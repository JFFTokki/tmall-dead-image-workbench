#!/usr/bin/env python3
"""
从 Excel 嵌入图片中识别疑似“死图”状态条。

核心思路：
1. 读取 xlsx 中每张嵌入图片的锚点位置；
2. 用锚点所在行匹配商品ID，用锚点所在列匹配图片列名；
3. 将图片写入临时目录用于检测，程序结束后自动清理；
4. 在图片底部区域扫描“横向状态标识”：
   - 位置必须靠近底部；
   - 形态应为横向矩形；
   - 顶部应有明显边界；
   - 与主体图片区背景应有明显差异；
   - 条内颜色应相对统一；
5. 只输出最终检测表格。

依赖：
    pip install openpyxl pillow numpy pandas

示例：
    python detect_dead_images.py export_469_20260706_135143_2932072.xlsx --output-dir output
"""

from __future__ import annotations

import argparse
import base64
import os
import gc
import json
import posixpath
import re
import tempfile
import zipfile
import urllib.error
import urllib.request
from dataclasses import dataclass
from io import BytesIO
from pathlib import Path
from typing import Iterable
import xml.etree.ElementTree as ET

import numpy as np
import openpyxl
import pandas as pd
from PIL import Image
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


DEFAULT_IMAGE_COLUMNS = [
    "1:1主图1",
    "1:1主图2",
    "3:4主图1",
    "3:4主图2",
]


JUDGMENT_HEADER = "死图判断"


@dataclass(frozen=True)
class DetectionConfig:
    """检测参数集中放在这里，方便按你的图片样式微调。"""

    # 只扫描图片底部这段区域。状态条如果不在底部，不应被判为死图。
    scan_start_frac: float = 0.55

    # 状态条的最小/最大高度占比。太薄通常是普通边缘，太高通常是主体背景。
    min_bar_height_frac: float = 0.035
    max_bar_height_frac: float = 0.32

    # 计算颜色特征时裁掉左右边缘，避免边框、压缩毛边影响判断。
    side_margin_frac: float = 0.04

    # 边界/背景对比的软阈值。不是硬过滤，主要用于打分。
    strong_boundary_rgb_diff: float = 28.0
    strong_background_rgb_diff: float = 32.0

    # 条内颜色越统一，分数越高。含文字的状态条允许有一定波动。
    good_uniformity_std: float = 30.0
    max_uniformity_std: float = 58.0

    # 分类阈值。
    high_confidence_threshold: float = 0.72
    review_threshold: float = 0.45

    # 模板图库命中阈值。相似度来自当前特征空间，不是百分比。
    template_strong_similarity: float = 0.40
    template_soft_similarity: float = 0.34
    template_weak_similarity: float = 0.28
    feedback_block_similarity: float = 0.96

    # debug 图中框的颜色和宽度。
    debug_box_color: tuple[int, int, int] = (255, 0, 0)
    debug_box_width: int = 3


@dataclass
class DetectionResult:
    is_suspected_dead: bool
    confidence: float
    reasons: list[str]
    box: tuple[int, int, int, int] | None
    category: str
    needs_review: bool = False
    review_reasons: list[str] | None = None




@dataclass(frozen=True)
class LlmConfig:
    enabled: bool
    api_url: str
    api_key: str
    model: str
    timeout_seconds: int = 60

@dataclass
class WorkbookImage:
    """从 xlsx drawing XML 中解析出的图片和锚点。"""

    row: int
    col: int
    media_path: str
    data: bytes
    extension: str


@dataclass
class LearningModel:
    """用人工标注样本训练出的本地轻量分类器。"""

    mean: np.ndarray
    std: np.ndarray
    weights: np.ndarray
    bias: float
    positive_vectors: np.ndarray
    negative_vectors: np.ndarray
    template_reference_vectors: np.ndarray
    feedback_negative_vectors: np.ndarray
    labeled_count: int
    positive_count: int
    negative_count: int
    template_reference_count: int

    def probability(self, vector: np.ndarray) -> float:
        z = (vector - self.mean) / self.std
        linear_score = float(np.dot(z, self.weights) + self.bias)
        logistic_prob = float(1.0 / (1.0 + np.exp(-np.clip(linear_score, -30.0, 30.0))))

        pos_dist = np.min(np.linalg.norm(self.positive_vectors - z, axis=1))
        neg_dist = np.min(np.linalg.norm(self.negative_vectors - z, axis=1))
        neighbor_prob = float(neg_dist / (pos_dist + neg_dist + 1e-6))

        # 逻辑回归负责总体边界，最近邻负责贴近人工样本。
        return float(0.65 * logistic_prob + 0.35 * neighbor_prob)

    def template_similarity(self, vector: np.ndarray) -> float:
        """返回与死图模板图库最相似的程度，1 表示几乎一致。"""

        if len(self.template_reference_vectors) == 0:
            return 0.0
        z = (vector - self.mean) / self.std
        min_dist = float(np.min(np.linalg.norm(self.template_reference_vectors - z, axis=1)))
        return float(np.exp(-min_dist / np.sqrt(len(z))))

    def feedback_negative_similarity(self, vector: np.ndarray) -> float:
        """返回与历史标黄误判样本的相似程度，1 表示几乎一致。"""

        if len(self.feedback_negative_vectors) == 0:
            return 0.0
        z = (vector - self.mean) / self.std
        min_dist = float(np.min(np.linalg.norm(self.feedback_negative_vectors - z, axis=1)))
        return float(np.exp(-min_dist / np.sqrt(len(z))))


def safe_filename(value: object, fallback: str = "unknown") -> str:
    """把商品ID/列名转成安全文件名，避免斜杠、冒号等字符影响保存。"""

    text = str(value).strip() if value is not None else fallback
    text = text or fallback
    text = re.sub(r"[\\/:*?\"<>|\s]+", "_", text)
    return text[:120]


def image_extension(image_format: str | None) -> str:
    fmt = (image_format or "png").lower()
    if fmt in {"jpeg", "jpg"}:
        return "jpg"
    if fmt in {"png", "gif", "bmp", "tiff", "webp"}:
        return fmt
    return "png"


def anchor_cell(image) -> tuple[int, int]:
    """返回图片锚点所在单元格，行列均为 1-based。"""

    marker = image.anchor._from
    return marker.row + 1, marker.col + 1


def extract_image_bytes(image) -> bytes:
    """openpyxl 的 Image 对象内部保存了原始图片字节。"""

    data = image._data()
    if isinstance(data, bytes):
        return data
    return bytes(data)


def _relationship_map(zip_file: zipfile.ZipFile, rels_path: str) -> dict[str, str]:
    rel_ns = {"pr": "http://schemas.openxmlformats.org/package/2006/relationships"}
    root = ET.fromstring(zip_file.read(rels_path))
    return {rel.attrib["Id"]: rel.attrib["Target"] for rel in root.findall("pr:Relationship", rel_ns)}


def _resolve_zip_target(base_file: str, target: str) -> str:
    """把关系文件里的相对路径解析成 zip 内部路径。"""

    if target.startswith("/"):
        return target.lstrip("/")
    base_dir = posixpath.dirname(base_file)
    return posixpath.normpath(posixpath.join(base_dir, target))


def _sheet_xml_path(zip_file: zipfile.ZipFile, sheet_name: str | None) -> str:
    """根据工作表名找到对应 worksheet XML 路径。"""

    main_ns = {"m": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
    rel_ns_uri = "{http://schemas.openxmlformats.org/officeDocument/2006/relationships}"
    workbook_root = ET.fromstring(zip_file.read("xl/workbook.xml"))
    workbook_rels = _relationship_map(zip_file, "xl/_rels/workbook.xml.rels")

    sheets = workbook_root.find("m:sheets", main_ns)
    if sheets is None:
        raise ValueError("工作簿中找不到 sheets 信息")

    for sheet in sheets.findall("m:sheet", main_ns):
        current_name = sheet.attrib.get("name")
        if sheet_name is None or current_name == sheet_name:
            rid = sheet.attrib[f"{rel_ns_uri}id"]
            return _resolve_zip_target("xl/workbook.xml", workbook_rels[rid])

    raise ValueError(f"找不到工作表：{sheet_name}")


def extract_workbook_images(input_xlsx: Path, sheet_name: str | None = None) -> list[WorkbookImage]:
    """
    直接解析 xlsx 内部 drawing XML 提取图片。

    这样每张图片都来自 Excel 文件自身记录的：
    - 锚点行列；
    - drawing relationship id；
    - 对应 xl/media 图片文件。

    比依赖 openpyxl 的内部 `_images` 顺序更稳，尤其适合复核“图片和商品ID是否错位”。
    """

    xdr_ns = {
        "xdr": "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing",
        "a": "http://schemas.openxmlformats.org/drawingml/2006/main",
    }
    rel_embed = "{http://schemas.openxmlformats.org/officeDocument/2006/relationships}embed"
    sheet_ns = {"m": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
    rel_ns_uri = "{http://schemas.openxmlformats.org/officeDocument/2006/relationships}"

    with zipfile.ZipFile(input_xlsx) as zip_file:
        sheet_xml = _sheet_xml_path(zip_file, sheet_name)
        sheet_rels_path = posixpath.join(posixpath.dirname(sheet_xml), "_rels", posixpath.basename(sheet_xml) + ".rels")
        if sheet_rels_path not in zip_file.namelist():
            return []

        sheet_root = ET.fromstring(zip_file.read(sheet_xml))
        drawing = sheet_root.find("m:drawing", sheet_ns)
        if drawing is None:
            return []

        sheet_rels = _relationship_map(zip_file, sheet_rels_path)
        drawing_rid = drawing.attrib[f"{rel_ns_uri}id"]
        drawing_xml = _resolve_zip_target(sheet_xml, sheet_rels[drawing_rid])
        drawing_rels_path = posixpath.join(
            posixpath.dirname(drawing_xml),
            "_rels",
            posixpath.basename(drawing_xml) + ".rels",
        )
        drawing_rels = _relationship_map(zip_file, drawing_rels_path)
        drawing_root = ET.fromstring(zip_file.read(drawing_xml))

        images: list[WorkbookImage] = []
        for anchor in list(drawing_root):
            tag_name = anchor.tag.split("}", 1)[-1]
            if tag_name not in {"oneCellAnchor", "twoCellAnchor"}:
                continue
            marker = anchor.find("xdr:from", xdr_ns)
            blip = anchor.find(".//a:blip", xdr_ns)
            if marker is None or blip is None:
                continue
            rid = blip.attrib.get(rel_embed)
            if not rid or rid not in drawing_rels:
                continue

            row = int(marker.find("xdr:row", xdr_ns).text) + 1
            col = int(marker.find("xdr:col", xdr_ns).text) + 1
            media_path = _resolve_zip_target(drawing_xml, drawing_rels[rid])
            extension = Path(media_path).suffix.lower().lstrip(".") or "png"
            if extension == "jpeg":
                extension = "jpg"
            images.append(
                WorkbookImage(
                    row=row,
                    col=col,
                    media_path=media_path,
                    data=zip_file.read(media_path),
                    extension=extension,
                )
            )

        return images


def rgb_distance(a: np.ndarray, b: np.ndarray) -> float:
    """RGB 均值之间的欧氏距离。"""

    return float(np.linalg.norm(a.astype(np.float32) - b.astype(np.float32)))


def clamp01(value: float) -> float:
    return max(0.0, min(1.0, value))


def detect_bottom_status_bar(image_path: Path, config: DetectionConfig) -> DetectionResult:
    """
    判断图片底部是否存在人工横向状态条。

    这里不按“绿色/红色”等固定颜色判断，而是扫描多个候选顶部边界，
    对每个候选状态条综合计算：
    - 是否靠底；
    - 高度是否像横条；
    - 顶部边界是否突变；
    - 条内颜色是否统一；
    - 条与上方主体区域是否有明显差异。
    """

    with Image.open(image_path) as img:
        rgb = img.convert("RGB")
        arr = np.asarray(rgb).astype(np.float32)

    height, width, _ = arr.shape
    if width < 20 or height < 20:
        return DetectionResult(False, 0.0, ["图片尺寸过小，无法可靠判断"], None, "review")

    x0 = int(width * config.side_margin_frac)
    x1 = max(x0 + 1, int(width * (1.0 - config.side_margin_frac)))
    cropped = arr[:, x0:x1, :]

    scan_start = int(height * config.scan_start_frac)
    min_bar_h = max(3, int(height * config.min_bar_height_frac))
    max_bar_h = max(min_bar_h + 1, int(height * config.max_bar_height_frac))
    boundary_band_h = max(2, min(8, height // 20))

    best: dict[str, object] | None = None

    # 只考虑“从某个顶部边界一直延伸到图片底部”的横条。
    # 如果你的状态条底部留有白边，可把 bottom_y 改成 height - 少量像素。
    for top_y in range(scan_start, height - min_bar_h):
        bar_h = height - top_y
        if bar_h > max_bar_h:
            continue

        above_start = max(0, top_y - max(boundary_band_h * 3, min_bar_h))
        above_end = max(above_start + 1, top_y)
        boundary_start = max(0, top_y - boundary_band_h)
        boundary_end = min(height, top_y + boundary_band_h)

        above = cropped[above_start:above_end]
        bar = cropped[top_y:height]
        boundary_above = cropped[boundary_start:top_y]
        boundary_below = cropped[top_y:boundary_end]
        if bar.size == 0 or boundary_above.size == 0 or boundary_below.size == 0:
            continue

        above_mean = above.reshape(-1, 3).mean(axis=0)
        bar_pixels = bar.reshape(-1, 3)
        bar_mean = bar_pixels.mean(axis=0)

        background_diff = rgb_distance(bar_mean, above_mean)
        boundary_diff = rgb_distance(
            boundary_above.reshape(-1, 3).mean(axis=0),
            boundary_below.reshape(-1, 3).mean(axis=0),
        )

        # 条内统一度：用每个像素到中位颜色的距离，避免少量文字/噪声过度影响。
        bar_median = np.median(bar_pixels, axis=0)
        pixel_deviation = np.linalg.norm(bar_pixels - bar_median, axis=1)
        uniformity_std = float(np.percentile(pixel_deviation, 75))

        # 横向矩形感：每一行的颜色均值变化较小，说明它像一整条横带。
        row_means = bar.mean(axis=1)
        row_variation = float(np.mean(np.linalg.norm(row_means - row_means.mean(axis=0), axis=1)))

        # 整条色块覆盖度：状态标识通常横跨底部大部分宽度；
        # 商品本身、角标或局部装饰即使有红/蓝色，也不应获得高分。
        bar_col_medians = np.median(bar, axis=0)
        global_bar_median = np.median(bar_col_medians, axis=0)
        column_deviation = np.linalg.norm(bar_col_medians - global_bar_median, axis=1)
        full_width_uniformity = float((column_deviation < 45.0).mean())

        boundary_above_cols = np.median(boundary_above, axis=0)
        boundary_below_cols = np.median(boundary_below, axis=0)
        boundary_col_diff = np.linalg.norm(boundary_above_cols - boundary_below_cols, axis=1)
        boundary_width_coverage = float((boundary_col_diff > 35.0).mean())

        position_score = clamp01((top_y / height - config.scan_start_frac) / (1.0 - config.scan_start_frac))
        height_score = clamp01(1.0 - abs((bar_h / height) - 0.12) / 0.18)
        boundary_score = clamp01(boundary_diff / config.strong_boundary_rgb_diff)
        background_score = clamp01(background_diff / config.strong_background_rgb_diff)
        uniformity_score = clamp01(
            (config.max_uniformity_std - uniformity_std)
            / (config.max_uniformity_std - config.good_uniformity_std)
        )
        rectangle_score = clamp01(1.0 - row_variation / 28.0)

        confidence = (
            0.12 * position_score
            + 0.10 * height_score
            + 0.18 * boundary_score
            + 0.16 * background_score
            + 0.14 * uniformity_score
            + 0.08 * rectangle_score
            + 0.12 * full_width_uniformity
            + 0.10 * boundary_width_coverage
        )

        # 关键约束：真正的状态标识应在底部形成一条横向边界。
        # 如果只有局部文字、人物脚部、角标或背景色块变化，边界覆盖度会偏低；
        # 这种情况即使有明显颜色差，也不应被当成死图。
        if boundary_width_coverage < 0.45:
            confidence *= 0.62
        elif boundary_width_coverage < 0.60:
            confidence *= 0.78
        if full_width_uniformity < 0.55:
            confidence *= 0.72
        if uniformity_std > 75.0:
            confidence *= 0.86

        candidate = {
            "top_y": top_y,
            "bar_h": bar_h,
            "confidence": confidence,
            "boundary_diff": boundary_diff,
            "background_diff": background_diff,
            "uniformity_std": uniformity_std,
            "row_variation": row_variation,
            "full_width_uniformity": full_width_uniformity,
            "boundary_width_coverage": boundary_width_coverage,
            "position_score": position_score,
            "height_score": height_score,
        }
        if best is None or confidence > float(best["confidence"]):
            best = candidate

    if best is None:
        return DetectionResult(False, 0.0, ["未找到可评估的底部候选横条"], None, "normal")

    confidence = round(float(best["confidence"]), 4)
    top_y = int(best["top_y"])
    box = (0, top_y, width - 1, height - 1)
    reasons = [
        f"底部候选条高度占比 {float(best['bar_h']) / height:.1%}",
        f"顶部边界差异 {float(best['boundary_diff']):.1f}",
        f"主体背景差异 {float(best['background_diff']):.1f}",
        f"条内颜色波动 {float(best['uniformity_std']):.1f}",
        f"横向行波动 {float(best['row_variation']):.1f}",
        f"整条覆盖度 {float(best['full_width_uniformity']):.2f}",
        f"边界覆盖度 {float(best['boundary_width_coverage']):.2f}",
    ]

    if confidence >= config.high_confidence_threshold:
        category = "high_confidence_dead"
        suspected = True
    elif confidence >= config.review_threshold:
        category = "review"
        suspected = True
    else:
        category = "normal"
        suspected = False

    return DetectionResult(suspected, confidence, reasons, box if suspected else None, category)


def bottom_visual_vector(image_path: Path) -> np.ndarray:
    """
    提取底部区域的视觉特征，用于和人工标注样本做相似度校准。

    特征只来自图片下半部分，避免商品主体差异主导判断；同时包含 RGB、
    饱和度、亮度和垂直边缘，能区分“状态条”和普通商品模板底栏。
    """

    with Image.open(image_path) as img:
        rgb = img.convert("RGB")
        width, height = rgb.size
        crop = rgb.crop((0, int(height * 0.45), width, height)).resize((32, 18), Image.Resampling.BILINEAR)

    arr = np.asarray(crop).astype(np.float32) / 255.0
    max_channel = arr.max(axis=2)
    min_channel = arr.min(axis=2)
    saturation = np.zeros_like(max_channel)
    np.divide(max_channel - min_channel, max_channel, out=saturation, where=max_channel > 1e-6)
    saturation = saturation[..., None]
    value = max_channel[..., None]

    gray = arr.mean(axis=2)
    vertical_edge = np.zeros_like(gray)
    vertical_edge[1:] = np.abs(gray[1:] - gray[:-1])

    return np.concatenate([arr, saturation, value, vertical_edge[..., None]], axis=2).reshape(-1)


def status_bar_feature_vector(image_path: Path, config: DetectionConfig) -> np.ndarray:
    """
    提取适合学习的低维特征。

    这些特征刻意描述“底部是否有人工状态条”的结构，而不是记住某种颜色：
    - 候选底部条的高度、位置；
    - 条内统一度、横向覆盖度；
    - 条和上方主体区域的整体差异；
    - 横向边界是否连续；
    - 底部区域的饱和度、亮度、边缘强度。
    """

    with Image.open(image_path) as img:
        arr = np.asarray(img.convert("RGB")).astype(np.float32)

    height, width, _ = arr.shape
    x0 = int(width * config.side_margin_frac)
    x1 = max(x0 + 1, int(width * (1.0 - config.side_margin_frac)))
    cropped = arr[:, x0:x1, :]
    scan_start = int(height * config.scan_start_frac)
    min_bar_h = max(3, int(height * config.min_bar_height_frac))
    max_bar_h = max(min_bar_h + 1, int(height * config.max_bar_height_frac))
    boundary_band_h = max(2, min(8, height // 20))

    best_features: np.ndarray | None = None
    best_score = -1.0

    for top_y in range(scan_start, height - min_bar_h):
        bar_h = height - top_y
        if bar_h > max_bar_h:
            continue

        above_start = max(0, top_y - max(boundary_band_h * 3, min_bar_h))
        above_end = max(above_start + 1, top_y)
        boundary_start = max(0, top_y - boundary_band_h)
        boundary_end = min(height, top_y + boundary_band_h)

        above = cropped[above_start:above_end]
        bar = cropped[top_y:height]
        boundary_above = cropped[boundary_start:top_y]
        boundary_below = cropped[top_y:boundary_end]
        if above.size == 0 or bar.size == 0 or boundary_above.size == 0 or boundary_below.size == 0:
            continue

        above_pixels = above.reshape(-1, 3)
        bar_pixels = bar.reshape(-1, 3)
        above_mean = above_pixels.mean(axis=0)
        bar_mean = bar_pixels.mean(axis=0)
        background_diff = rgb_distance(bar_mean, above_mean)
        boundary_diff = rgb_distance(
            boundary_above.reshape(-1, 3).mean(axis=0),
            boundary_below.reshape(-1, 3).mean(axis=0),
        )

        bar_median = np.median(bar_pixels, axis=0)
        pixel_deviation = np.linalg.norm(bar_pixels - bar_median, axis=1)
        uniformity_p75 = float(np.percentile(pixel_deviation, 75))
        uniformity_p90 = float(np.percentile(pixel_deviation, 90))

        row_means = bar.mean(axis=1)
        row_variation = float(np.mean(np.linalg.norm(row_means - row_means.mean(axis=0), axis=1)))

        bar_col_medians = np.median(bar, axis=0)
        global_bar_median = np.median(bar_col_medians, axis=0)
        column_deviation = np.linalg.norm(bar_col_medians - global_bar_median, axis=1)
        full_width_uniformity = float((column_deviation < 45.0).mean())

        boundary_above_cols = np.median(boundary_above, axis=0)
        boundary_below_cols = np.median(boundary_below, axis=0)
        boundary_col_diff = np.linalg.norm(boundary_above_cols - boundary_below_cols, axis=1)
        boundary_width_coverage = float((boundary_col_diff > 35.0).mean())
        boundary_p25 = float(np.percentile(boundary_col_diff, 25))
        boundary_p50 = float(np.percentile(boundary_col_diff, 50))

        max_channel = bar_pixels.max(axis=1)
        min_channel = bar_pixels.min(axis=1)
        saturation = np.zeros_like(max_channel)
        np.divide(max_channel - min_channel, max_channel, out=saturation, where=max_channel > 1e-6)
        value = max_channel / 255.0
        saturated_coverage = float((saturation > 0.22).mean())
        near_white_coverage = float(((min_channel > 230.0) & (saturation < 0.08)).mean())

        # 一个只用于挑“最像底部条”的候选，不直接作为最终分类。
        candidate_score = (
            0.22 * boundary_width_coverage
            + 0.20 * full_width_uniformity
            + 0.18 * clamp01(boundary_diff / 80.0)
            + 0.16 * clamp01(background_diff / 90.0)
            + 0.12 * clamp01((120.0 - uniformity_p75) / 120.0)
            + 0.12 * clamp01(1.0 - row_variation / 50.0)
        )

        feature_values = np.array(
            [
                top_y / height,
                bar_h / height,
                clamp01(boundary_diff / 160.0),
                clamp01(background_diff / 180.0),
                clamp01(uniformity_p75 / 255.0),
                clamp01(uniformity_p90 / 255.0),
                clamp01(row_variation / 80.0),
                full_width_uniformity,
                boundary_width_coverage,
                clamp01(boundary_p25 / 120.0),
                clamp01(boundary_p50 / 140.0),
                float(np.mean(saturation)),
                float(np.median(saturation)),
                saturated_coverage,
                float(np.mean(value)),
                float(np.std(value)),
                clamp01(candidate_score),
                near_white_coverage,
            ],
            dtype=np.float32,
        )

        if candidate_score > best_score:
            best_score = candidate_score
            best_features = feature_values

    if best_features is None:
        return np.zeros(18, dtype=np.float32)
    return best_features


def learning_vector(image_path: Path, config: DetectionConfig) -> np.ndarray:
    """组合结构特征和压缩后的底部视觉特征。"""

    visual = bottom_visual_vector(image_path)
    # 原始底部像素维度较高，取固定分位摘要，减少过拟合。
    visual_summary = np.array(
        [
            float(np.mean(visual)),
            float(np.std(visual)),
            float(np.percentile(visual, 10)),
            float(np.percentile(visual, 25)),
            float(np.percentile(visual, 50)),
            float(np.percentile(visual, 75)),
            float(np.percentile(visual, 90)),
        ],
        dtype=np.float32,
    )
    return np.concatenate([status_bar_feature_vector(image_path, config), visual_summary])


def build_learning_model(
    vectors: list[np.ndarray],
    labels: list[int],
    template_reference_count: int = 0,
    feedback_sample_count: int = 0,
) -> LearningModel | None:
    """从人工标注样本中训练一个轻量逻辑回归分类器。"""

    if not vectors or 1 not in labels or 0 not in labels:
        return None
    matrix = np.vstack(vectors)
    mean = matrix.mean(axis=0)
    std = matrix.std(axis=0) + 1e-6
    normalized = np.clip((matrix - mean) / std, -6.0, 6.0)
    label_array = np.asarray(labels)

    weights = np.zeros(normalized.shape[1], dtype=np.float32)
    bias = 0.0
    pos_weight = len(label_array) / (2.0 * max(1, int((label_array == 1).sum())))
    neg_weight = len(label_array) / (2.0 * max(1, int((label_array == 0).sum())))
    sample_weights = np.where(label_array == 1, pos_weight, neg_weight).astype(np.float32)

    learning_rate = 0.08
    l2 = 0.015
    y = label_array.astype(np.float32)
    for _ in range(1800):
        logits = np.clip(normalized @ weights + bias, -30.0, 30.0)
        pred = 1.0 / (1.0 + np.exp(-logits))
        error = (pred - y) * sample_weights
        grad_w = (normalized.T @ error) / len(y) + l2 * weights
        grad_b = float(error.mean())
        weights -= learning_rate * grad_w
        bias -= learning_rate * grad_b

    feedback_start = len(labels) - template_reference_count - feedback_sample_count
    feedback_end = feedback_start + feedback_sample_count
    feedback_vectors = normalized[feedback_start:feedback_end] if feedback_sample_count else np.empty((0, normalized.shape[1]))
    feedback_label_array = label_array[feedback_start:feedback_end] if feedback_sample_count else np.asarray([])
    return LearningModel(
        mean=mean,
        std=std,
        weights=weights,
        bias=bias,
        positive_vectors=normalized[label_array == 1],
        negative_vectors=normalized[label_array == 0],
        template_reference_vectors=normalized[-template_reference_count:] if template_reference_count else np.empty((0, normalized.shape[1])),
        feedback_negative_vectors=feedback_vectors[feedback_label_array == 0] if feedback_sample_count else np.empty((0, normalized.shape[1])),
        labeled_count=len(labels),
        positive_count=int((label_array == 1).sum()),
        negative_count=int((label_array == 0).sum()),
        template_reference_count=template_reference_count,
    )


def find_template_reference_dir(input_xlsx: Path, explicit_dir: Path | None) -> Path | None:
    """定位 template_reference；默认从当前目录和输入文件附近查找。"""

    if explicit_dir is not None:
        return explicit_dir if explicit_dir.exists() else None

    candidates = [
        Path.cwd() / "template_reference",
        input_xlsx.parent / "template_reference",
        input_xlsx.parent.parent / "template_reference",
        Path(__file__).resolve().parent.parent / "template_reference",
    ]
    for candidate in candidates:
        if candidate.exists():
            return candidate
    return None


def iter_template_reference_images(reference_dir: Path) -> list[Path]:
    """
    读取模板参考图。

    这些模板图案属于死图参考样式，作为正样本可让模型识别相似死图模板。
    """

    index_path = reference_dir / "templates_index.json"
    if index_path.exists():
        items = json.loads(index_path.read_text(encoding="utf-8"))
        seen_hashes: set[str] = set()
        paths: list[Path] = []
        for item in items:
            file_value = item.get("file")
            sha256 = item.get("sha256")
            if item.get("duplicateOf") or not file_value:
                continue
            if sha256 and sha256 in seen_hashes:
                continue
            if sha256:
                seen_hashes.add(sha256)
            image_path = reference_dir / str(file_value)
            if image_path.exists():
                paths.append(image_path)
        if paths:
            return paths

    raw_dir = reference_dir / "raw"
    if not raw_dir.exists():
        return []
    return sorted(
        path
        for path in raw_dir.iterdir()
        if path.suffix.lower() in {".png", ".jpg", ".jpeg", ".webp", ".bmp"}
    )


def read_template_reference_vectors(
    reference_dir: Path | None,
    config: DetectionConfig,
) -> list[np.ndarray]:
    """把模板参考图库转成死图正样本特征。"""

    if reference_dir is None:
        return []

    vectors: list[np.ndarray] = []
    for image_path in iter_template_reference_images(reference_dir):
        try:
            vectors.append(learning_vector(image_path, config))
        except Exception:
            # 单张模板损坏不应中断主流程；跳过即可。
            continue
    return vectors


def load_feedback_samples(feedback_path: Path) -> tuple[list[np.ndarray], list[int], set[str]]:
    """读取已落盘的人工反馈样本。"""

    if not feedback_path.exists():
        return [], [], set()

    data = json.loads(feedback_path.read_text(encoding="utf-8"))
    vectors: list[np.ndarray] = []
    labels: list[int] = []
    keys: set[str] = set()
    for sample in data.get("samples", []):
        vector = sample.get("vector")
        label = sample.get("label")
        key = sample.get("key")
        if not isinstance(vector, list) or label not in {0, 1} or not key:
            continue
        # v1 feedback vectors had 24 dims. Current vector adds near-white-bottom
        # coverage before the 7 visual summary features, so keep old feedback usable.
        if len(vector) == 24:
            vector = vector[:17] + [0.0] + vector[17:]
        vectors.append(np.asarray(vector, dtype=np.float32))
        labels.append(int(label))
        keys.add(str(key))
    return vectors, labels, keys


def save_feedback_samples(
    feedback_path: Path,
    existing_keys: set[str],
    new_samples: list[dict[str, object]],
) -> int:
    """把本次批注学习到的样本追加保存，按 key 去重。"""

    if feedback_path.exists():
        data = json.loads(feedback_path.read_text(encoding="utf-8"))
    else:
        data = {"version": 1, "samples": []}

    samples = data.setdefault("samples", [])
    added = 0
    for sample in new_samples:
        key = str(sample.get("key", ""))
        if not key or key in existing_keys:
            continue
        samples.append(sample)
        existing_keys.add(key)
        added += 1

    feedback_path.parent.mkdir(parents=True, exist_ok=True)
    feedback_path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
    return added


def calibrated_result(
    image_path: Path,
    config: DetectionConfig,
    learning_model: LearningModel | None,
    forced_label: int | None,
) -> DetectionResult:
    """综合结构检测、学习模型，以及人工标签覆盖。"""

    shape_result = detect_bottom_status_bar(image_path, config)

    if forced_label is not None:
        suspected = forced_label > 0
        confidence = 1.0 if suspected else 0.0
        category = "high_confidence_dead" if suspected else "normal"
        reasons = ["人工校准标签：死图" if suspected else "人工修正标签：正常"]
        return DetectionResult(suspected, confidence, reasons, shape_result.box if suspected else None, category)

    if learning_model is None:
        return shape_result

    vector = learning_vector(image_path, config)
    model_prob = learning_model.probability(vector)
    confidence = round(0.82 * model_prob + 0.18 * shape_result.confidence, 4)
    template_similarity = learning_model.template_similarity(vector)
    feedback_negative_similarity = learning_model.feedback_negative_similarity(vector)
    near_white_bottom = bool(len(vector) > 17 and vector[17] >= 0.82)
    template_boost = 0.0
    feedback_block = (
        feedback_negative_similarity >= config.feedback_block_similarity
        and feedback_negative_similarity >= template_similarity + 0.35
    )
    has_bottom_full_width_bar = (
        shape_result.confidence >= config.review_threshold
        and shape_result.box is not None
        and (shape_result.box[1] / max(1, shape_result.box[3] + 1)) >= config.scan_start_frac
        and any(reason.startswith("整条覆盖度 1.00") for reason in shape_result.reasons)
        and any(reason.startswith("边界覆盖度 0.7") or reason.startswith("边界覆盖度 0.8") or reason.startswith("边界覆盖度 0.9") or reason.startswith("边界覆盖度 1.0") for reason in shape_result.reasons)
    )
    if feedback_block:
        confidence = min(confidence, 0.18)
    elif has_bottom_full_width_bar and not near_white_bottom and template_similarity >= config.template_strong_similarity:
        template_boost = 1.0
    elif has_bottom_full_width_bar and not near_white_bottom and template_similarity >= config.template_soft_similarity:
        template_boost = 1.0
    elif has_bottom_full_width_bar and not near_white_bottom and template_similarity >= config.template_weak_similarity:
        template_boost = 1.0
    confidence = round(max(confidence, template_boost), 4)
    if near_white_bottom and not has_bottom_full_width_bar:
        confidence = min(confidence, 0.18)

    reasons = shape_result.reasons + [
        f"学习模型概率 {model_prob:.3f}",
        f"死图模板相似度 {template_similarity:.3f}",
        f"历史误判相似度 {feedback_negative_similarity:.3f}",
        f"底部贯穿长条 {'是' if has_bottom_full_width_bar else '否'}",
        f"底部近白留白 {'是' if near_white_bottom else '否'}",
        f"模板命中提升 {template_boost:.2f}",
        (
            f"训练样本 {learning_model.labeled_count}（死图 {learning_model.positive_count}，"
            f"正常 {learning_model.negative_count}，模板参考 {learning_model.template_reference_count}）"
        ),
    ]

    if confidence >= config.high_confidence_threshold:
        category = "high_confidence_dead"
        suspected = True
    elif confidence >= config.review_threshold:
        category = "review"
        suspected = True
    else:
        category = "normal"
        suspected = False

    review_reasons: list[str] = []
    if category == "review":
        review_reasons.append("达到疑似阈值但未达到高置信阈值")
    if suspected and not has_bottom_full_width_bar:
        review_reasons.append("缺少底部贯穿长条结构")
    if template_similarity >= config.template_weak_similarity and feedback_negative_similarity >= config.template_weak_similarity:
        review_reasons.append("同时接近死图模板和历史误判样本")
    if 0.82 <= confidence < config.high_confidence_threshold:
        review_reasons.append("置信度接近高置信边界")

    return DetectionResult(
        suspected,
        confidence,
        reasons,
        shape_result.box if suspected else None,
        category,
        needs_review=bool(review_reasons),
        review_reasons=review_reasons,
    )




def parse_llm_response(content: object) -> dict[str, object]:
    """Parse JSON returned by a vision model, tolerating markdown fences."""

    if isinstance(content, dict):
        raw = content
    else:
        text = str(content or "").strip()
        fence_match = re.search(r"```(?:json)?\s*(.*?)\s*```", text, flags=re.IGNORECASE | re.DOTALL)
        if fence_match:
            text = fence_match.group(1).strip()
        json_match = re.search(r"\{.*\}", text, flags=re.DOTALL)
        if json_match:
            text = json_match.group(0)
        raw = json.loads(text)

    is_dead = bool(raw.get("is_dead_image", raw.get("dead_image", raw.get("is_dead", False))))
    try:
        confidence = float(raw.get("confidence", 0.0))
    except (TypeError, ValueError):
        confidence = 0.0
    confidence = max(0.0, min(1.0, confidence))
    reason = str(raw.get("reason", raw.get("rationale", ""))).strip()
    return {"is_dead_image": is_dead, "confidence": confidence, "reason": reason}


def merge_llm_result(local_result: DetectionResult, llm_result: dict[str, object]) -> DetectionResult:
    """Merge model judgment without lowering recall from the local detector."""

    is_dead = bool(llm_result.get("is_dead_image"))
    confidence = max(0.0, min(1.0, float(llm_result.get("confidence") or 0.0)))
    reason = str(llm_result.get("reason") or "").strip() or "no reason returned"
    label = "\u7591\u4f3c\u6b7b\u56fe" if is_dead else "\u6b63\u5e38"
    reasons = list(local_result.reasons)
    review_reasons = list(local_result.review_reasons or [])
    reasons.append(f"\u5927\u6a21\u578b\u5224\u65ad\uff1a{label}\uff0c\u7f6e\u4fe1\u5ea6 {confidence:.2f}\uff0c\u539f\u56e0\uff1a{reason}")

    if is_dead:
        merged_confidence = max(local_result.confidence, confidence)
        category = "high_confidence_dead" if confidence >= 0.75 else "review"
        if local_result.category == "high_confidence_dead":
            category = "high_confidence_dead"
        if category == "review":
            review_reasons.append("\u5927\u6a21\u578b\u8ba4\u4e3a\u7591\u4f3c\u6b7b\u56fe\u4f46\u7f6e\u4fe1\u5ea6\u672a\u8fbe\u5230\u9ad8\u7f6e\u4fe1\u9608\u503c")
        return DetectionResult(
            is_suspected_dead=True,
            confidence=round(merged_confidence, 4),
            reasons=reasons,
            box=local_result.box,
            category=category,
            needs_review=bool(review_reasons) or category == "review",
            review_reasons=review_reasons,
        )

    if local_result.is_suspected_dead:
        review_reasons.append(f"\u5927\u6a21\u578b\u8ba4\u4e3a\u6b63\u5e38\uff0c\u6309\u53ec\u56de\u4f18\u5148\u4fdd\u7559\u672c\u5730\u7591\u4f3c\u7ed3\u679c\uff1a{reason}")
        return DetectionResult(
            is_suspected_dead=True,
            confidence=local_result.confidence,
            reasons=reasons,
            box=local_result.box,
            category=local_result.category,
            needs_review=True,
            review_reasons=review_reasons,
        )

    return DetectionResult(
        is_suspected_dead=False,
        confidence=local_result.confidence,
        reasons=reasons,
        box=local_result.box,
        category=local_result.category,
        needs_review=local_result.needs_review,
        review_reasons=review_reasons,
    )


def should_call_llm(result: DetectionResult) -> bool:
    return result.category == "review"


def call_multimodal_llm(image_path: Path, config: LlmConfig) -> dict[str, object]:
    suffix = image_path.suffix.lower().lstrip(".") or "png"
    mime = "image/jpeg" if suffix in {"jpg", "jpeg"} else f"image/{suffix}"
    encoded = base64.b64encode(image_path.read_bytes()).decode("ascii")
    prompt = (
        "\u4f60\u662f\u7535\u5546\u5546\u54c1\u56fe\u8d28\u68c0\u52a9\u624b\u3002"
        "\u8bf7\u5224\u65ad\u8fd9\u5f20\u56fe\u7247\u662f\u5426\u662f\u6b7b\u56fe\u6216\u5f02\u5e38\u56fe\u3002"
        "\u91cd\u70b9\u89c2\u5bdf\u56fe\u7247\u5e95\u90e8\u662f\u5426\u5b58\u5728\u4e0e\u4e3b\u4f53\u4e0d\u4e00\u81f4\u7684\u6a2a\u6761\u3001"
        "\u72b6\u6001\u6761\u3001\u6570\u5b57\u6b8b\u7559\u3001\u52a0\u8f7d\u5931\u8d25\u75d5\u8ff9\u6216\u5f02\u5e38\u906e\u6321\u3002"
        "\u53ea\u8fd4\u56de JSON\uff0c\u4e0d\u8981\u89e3\u91ca\u989d\u5916\u6587\u5b57\u3002\u683c\u5f0f\uff1a"
        '{"is_dead_image": true/false, "confidence": 0.0, "reason": "short reason"}'
    )
    payload = {
        "model": config.model,
        "messages": [
            {
                "role": "user",
                "content": [
                    {"type": "text", "text": prompt},
                    {"type": "image_url", "image_url": {"url": f"data:{mime};base64,{encoded}"}},
                ],
            }
        ],
        "temperature": 0,
        "response_format": {"type": "json_object"},
    }
    data = json.dumps(payload, ensure_ascii=False).encode("utf-8")
    request = urllib.request.Request(
        config.api_url,
        data=data,
        headers={"Authorization": f"Bearer {config.api_key}", "Content-Type": "application/json"},
        method="POST",
    )
    try:
        with urllib.request.urlopen(request, timeout=config.timeout_seconds) as response:
            body = response.read().decode("utf-8", errors="replace")
    except urllib.error.HTTPError as exc:
        detail = exc.read().decode("utf-8", errors="replace")[:500]
        raise RuntimeError(f"\u5927\u6a21\u578b\u63a5\u53e3 HTTP {exc.code}: {detail}") from exc
    raw = json.loads(body)
    content = raw.get("choices", [{}])[0].get("message", {}).get("content", raw)
    return parse_llm_response(content)


def llm_config_from_args(args: argparse.Namespace) -> LlmConfig | None:
    enabled = bool(getattr(args, "llm_enabled", False)) or os.environ.get("DEAD_IMAGE_LLM_ENABLED") == "1"
    if not enabled:
        return None
    api_url = (getattr(args, "llm_api_url", None) or os.environ.get("DEAD_IMAGE_LLM_API_URL") or "").strip()
    api_key = (getattr(args, "llm_api_key", None) or os.environ.get("DEAD_IMAGE_LLM_API_KEY") or "").strip()
    model = (getattr(args, "llm_model", None) or os.environ.get("DEAD_IMAGE_LLM_MODEL") or "").strip()
    timeout = int(getattr(args, "llm_timeout", None) or os.environ.get("DEAD_IMAGE_LLM_TIMEOUT") or 60)
    if not api_url or not api_key or not model:
        raise ValueError("\u542f\u7528\u5927\u6a21\u578b\u8f85\u52a9\u65f6\uff0c\u5fc5\u987b\u63d0\u4f9b API \u5730\u5740\u3001API Key \u548c\u6a21\u578b\u540d\u79f0")
    return LlmConfig(True, api_url, api_key, model, timeout_seconds=timeout)


def find_column(headers: dict[str, int], names: Iterable[str], label: str) -> int:
    for name in names:
        if name in headers:
            return headers[name]
    raise ValueError(f"找不到{label}列，候选列名：{', '.join(names)}")


def read_template_layout(
    template_xlsx: Path,
    image_column_names: list[str],
    calibration_rows: int,
) -> tuple[dict[tuple[int, str], int], dict[str, int]]:
    """
    读取示例输出文件：
    - 人工标签：前 N 行中，图片列右侧“死图判断”的 1/空白；
    - 输出布局：每个图片列对应哪个“死图判断”列。
    """

    wb = openpyxl.load_workbook(template_xlsx)
    ws = wb.active

    judgment_cols: dict[str, int] = {}
    manual_labels: dict[tuple[int, str], int] = {}
    for col in range(1, ws.max_column):
        header = ws.cell(1, col).value
        next_header = ws.cell(1, col + 1).value
        if header in image_column_names and next_header == "死图判断":
            image_col_name = str(header)
            judgment_cols[image_col_name] = col + 1

    for row_idx in range(2, min(ws.max_row, calibration_rows + 1) + 1):
        for image_col_name, label_col in judgment_cols.items():
            manual_labels[(row_idx, image_col_name)] = 1 if ws.cell(row_idx, label_col).value == 1 else 0

    return manual_labels, judgment_cols


def read_yellow_corrections(
    correction_xlsx: Path,
    image_column_names: list[str],
) -> tuple[dict[tuple[int, str], int], dict[str, int]]:
    """
    读取人工复核文件中的黄色标记。

    约定：在“死图判断”列中，被标黄的单元格表示原先误判，
    也就是对应图片应作为“正常图”负样本。
    """

    wb = openpyxl.load_workbook(correction_xlsx)
    ws = wb.active
    corrections: dict[tuple[int, str], int] = {}
    judgment_cols: dict[str, int] = {}

    for col in range(1, ws.max_column):
        header = ws.cell(1, col).value
        next_header = ws.cell(1, col + 1).value
        if header in image_column_names and next_header == "死图判断":
            judgment_cols[str(header)] = col + 1

    for image_col_name, label_col in judgment_cols.items():
        for row_idx in range(2, ws.max_row + 1):
            cell = ws.cell(row_idx, label_col)
            is_yellow = (
                cell.fill.fill_type not in (None, "none")
                and cell.fill.fgColor.type == "rgb"
                and cell.fill.fgColor.rgb in {"FFFFFF00", "FFFF00"}
            )
            if is_yellow:
                corrections[(row_idx, image_col_name)] = -1
            elif cell.value in {1, 2, "疑似"}:
                corrections[(row_idx, image_col_name)] = 1

    return corrections, judgment_cols


def write_template_output(
    template_xlsx: Path,
    output_xlsx: Path,
    results_by_cell: dict[tuple[int, str], DetectionResult],
    judgment_cols: dict[str, int],
) -> None:
    """按示例文件的版式输出：命中死图填 1，正常留空。"""

    wb = openpyxl.load_workbook(template_xlsx)
    ws = wb.active

    for row_idx in range(2, ws.max_row + 1):
        for image_col_name, label_col in judgment_cols.items():
            result = results_by_cell.get((row_idx, image_col_name))
            if result and result.category == "high_confidence_dead":
                ws.cell(row_idx, label_col).value = 1
            elif result and result.category == "review":
                ws.cell(row_idx, label_col).value = 1
            else:
                ws.cell(row_idx, label_col).value = None

    output_xlsx.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_xlsx)


def write_review_report(output_xlsx: Path, rows: list[dict[str, object]]) -> None:
    """输出二次审核明细，方便大文件先集中看风险项。"""

    if not rows:
        return
    review_path = output_xlsx.with_name(output_xlsx.stem + "_二次审核明细.xlsx")
    review_rows = [row for row in rows if row.get("需要二次审核") == "是" or row.get("处理状态") != "ok"]
    if not review_rows:
        review_rows = []
    pd.DataFrame(review_rows).to_excel(review_path, index=False)


def write_checkpoint(checkpoint_path: Path, rows: list[dict[str, object]]) -> None:
    """把阶段性判断明细保存为 JSONL，崩溃后至少能追溯已处理部分。"""

    checkpoint_path.parent.mkdir(parents=True, exist_ok=True)
    with checkpoint_path.open("w", encoding="utf-8") as file:
        for row in rows:
            file.write(json.dumps(row, ensure_ascii=False) + "\n")


def style_final_workbook(output_xlsx: Path) -> None:
    """
    美化最终结果表，方便人工复核。

    规则：
    - 冻结首行和商品ID列；
    - 图片列宽统一，判断列窄且居中；
    - 判断列中的 1 用红色高亮；
    - 商品ID避免科学计数法；
    - 保留图片，不做删除或重排。
    """

    wb = openpyxl.load_workbook(output_xlsx)
    ws = wb.active
    ws.freeze_panes = "B2"
    ws.auto_filter.ref = ws.dimensions
    ws.sheet_view.showGridLines = False

    header_fill = PatternFill("solid", fgColor="1F4E78")
    judge_header_fill = PatternFill("solid", fgColor="9C0006")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    judge_fill = PatternFill("solid", fgColor="FFC7CE")
    blank_fill = PatternFill("solid", fgColor="F7F7F7")
    judge_font = Font(color="9C0006", bold=True, size=12)
    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)

    judgment_cols: set[int] = set()
    for col_idx in range(1, ws.max_column + 1):
        header = ws.cell(1, col_idx).value
        column_letter = ws.cell(1, col_idx).column_letter
        if header == JUDGMENT_HEADER:
            judgment_cols.add(col_idx)
            ws.column_dimensions[column_letter].width = 9
        elif col_idx == 1:
            ws.column_dimensions[column_letter].width = 16
        else:
            ws.column_dimensions[column_letter].width = 11

        cell = ws.cell(1, col_idx)
        cell.fill = judge_header_fill if header == JUDGMENT_HEADER else header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = border

    for row_idx in range(2, ws.max_row + 1):
        ws.cell(row_idx, 1).number_format = "0"
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row_idx, col_idx)
            cell.border = border
            if col_idx == 1:
                cell.alignment = left
            else:
                cell.alignment = center
            if col_idx in judgment_cols:
                if cell.value == 1:
                    cell.fill = judge_fill
                    cell.font = judge_font
                else:
                    cell.fill = blank_fill
        if ws.max_row <= 12000:
            ws.row_dimensions[row_idx].height = max(ws.row_dimensions[row_idx].height or 0, 50)

    ws.row_dimensions[1].height = 24
    wb.save(output_xlsx)
    wb.close()


def process_workbook(
    input_xlsx: Path,
    sheet_name: str | None,
    id_column_name: str,
    image_column_names: list[str],
    config: DetectionConfig,
    manual_labels: dict[tuple[int, str], int] | None = None,
    template_reference_vectors: list[np.ndarray] | None = None,
    feedback_vectors: list[np.ndarray] | None = None,
    feedback_labels: list[int] | None = None,
    feedback_path: Path | None = None,
    feedback_existing_keys: set[str] | None = None,
    batch_size: int = 500,
    checkpoint_path: Path | None = None,
    llm_config: LlmConfig | None = None,
) -> pd.DataFrame:
    wb = openpyxl.load_workbook(input_xlsx)
    ws = wb[sheet_name] if sheet_name else wb.active

    headers = {}
    for col in range(1, ws.max_column + 1):
        value = ws.cell(1, col).value
        if value is not None:
            headers[str(value).strip()] = col

    id_col = find_column(headers, [id_column_name], "商品ID")
    image_cols = {name: headers[name] for name in image_column_names if name in headers}
    image_col_by_index = {col: name for name, col in image_cols.items()}

    if not image_cols:
        raise ValueError(f"找不到任何图片列：{', '.join(image_column_names)}")

    rows: list[dict[str, object]] = []
    training_vectors: list[np.ndarray] = []
    training_labels: list[int] = []
    new_feedback_samples: list[dict[str, object]] = []
    image_jobs: list[dict[str, object]] = []

    with tempfile.TemporaryDirectory(prefix="dead_image_detect_") as temp_dir:
        temp_path = Path(temp_dir)
        images = extract_workbook_images(input_xlsx, sheet_name=ws.title)
        for image_index, image in enumerate(images, start=1):
            row_idx, col_idx = image.row, image.col
            image_col_name = image_col_by_index.get(col_idx)

            # 只处理指定图片列、且不处理表头行上的图片。
            if row_idx <= 1 or image_col_name is None:
                continue

            product_id = ws.cell(row_idx, id_col).value
            ext = image.extension
            base_name = (
                f"row{row_idx:04d}_"
                f"{safe_filename(product_id, 'no_product_id')}_"
                f"{safe_filename(image_col_name)}_"
                f"img{image_index:04d}.{ext}"
            )
            image_path = temp_path / base_name
            image_path.write_bytes(image.data)

            manual_label = None if manual_labels is None else manual_labels.get((row_idx, image_col_name))
            if manual_label is not None:
                vector = learning_vector(image_path, config)
                label = 0 if manual_label == -1 else manual_label
                training_vectors.append(vector)
                training_labels.append(label)
                new_feedback_samples.append(
                    {
                        "key": f"{product_id}|{image_col_name}|{label}",
                        "label": label,
                        "product_id": str(product_id),
                        "image_col_name": image_col_name,
                        "source": "yellow_correction" if label == 0 else "confirmed_dead",
                        "vector": vector.tolist(),
                    }
                )

            # 示例空白只作为正常训练样本，不直接覆盖后续模板/模型判断；
            # 黄色修正(-1)和明确死图(1)才作为强制标签。
            forced_label = manual_label if manual_label in {-1, 1} else None

            image_jobs.append(
                {
                    "row_idx": row_idx,
                    "product_id": product_id,
                    "image_col_name": image_col_name,
                    "image_path": image_path,
                    "forced_label": forced_label,
                }
            )

        template_reference_count = len(template_reference_vectors or [])
        if feedback_vectors and feedback_labels:
            training_vectors.extend(feedback_vectors)
            training_labels.extend(feedback_labels)
        if template_reference_vectors:
            training_vectors.extend(template_reference_vectors)
            training_labels.extend([1] * len(template_reference_vectors))

        added_feedback_count = 0
        if feedback_path is not None and feedback_existing_keys is not None and new_feedback_samples:
            added_feedback_count = save_feedback_samples(
                feedback_path=feedback_path,
                existing_keys=feedback_existing_keys,
                new_samples=new_feedback_samples,
            )

        learning_model = build_learning_model(
            training_vectors,
            training_labels,
            template_reference_count=template_reference_count,
            feedback_sample_count=len(feedback_labels or []),
        )
        results_by_cell: dict[tuple[int, str], DetectionResult] = {}
        total_jobs = len(image_jobs)
        progress_counts = {"high_confidence_dead": 0, "review": 0, "normal": 0, "error": 0}
        print(f"开始判断图片：共 {total_jobs} 张；每 {max(1, batch_size)} 张保存一次 checkpoint", flush=True)

        for processed_count, job in enumerate(image_jobs, start=1):
            row_idx = int(job["row_idx"])
            product_id = job["product_id"]
            image_col_name = str(job["image_col_name"])
            image_path = Path(job["image_path"])
            forced_label = job["forced_label"]

            try:
                result = calibrated_result(
                    image_path=image_path,
                    config=config,
                    learning_model=learning_model,
                    forced_label=int(forced_label) if forced_label is not None else None,
                )
                llm_status = "???"
                llm_reason = ""
                if llm_config is not None and should_call_llm(result):
                    try:
                        llm_result = call_multimodal_llm(image_path, llm_config)
                        result = merge_llm_result(result, llm_result)
                        llm_status = "???"
                        llm_reason = str(llm_result.get("reason") or "")
                    except Exception as llm_exc:
                        llm_status = "????"
                        llm_reason = f"{type(llm_exc).__name__}: {llm_exc}"
                        result.review_reasons = list(result.review_reasons or []) + [f"????????{llm_reason}"]
                        result.needs_review = True
                status = "ok"
                error_message = ""
            except Exception as exc:
                result = DetectionResult(
                    is_suspected_dead=True,
                    confidence=0.55,
                    reasons=["单图处理异常，按召回优先进入二次审核"],
                    box=None,
                    category="review",
                    needs_review=True,
                    review_reasons=[f"处理异常：{type(exc).__name__}: {exc}"],
                )
                llm_status = "???"
                llm_reason = ""
                status = "error"
                error_message = f"{type(exc).__name__}: {exc}"

            results_by_cell[(row_idx, image_col_name)] = result
            progress_counts[result.category] = progress_counts.get(result.category, 0) + 1
            if status != "ok":
                progress_counts["error"] = progress_counts.get("error", 0) + 1

            rows.append(
                {
                    "Excel行号": row_idx,
                    "商品ID": product_id,
                    "图片列名": image_col_name,
                    "是否疑似死图": "是" if result.is_suspected_dead else "否",
                    "判断级别": result.category,
                    "置信度": result.confidence,
                    "命中原因": "；".join(result.reasons),
                    "需要二次审核": "是" if result.needs_review else "否",
                    "二次审核原因": "；".join(result.review_reasons or []),
                    "处理状态": status,
                    "错误信息": error_message,
                }
            )

            if checkpoint_path is not None and processed_count % max(1, batch_size) == 0:
                write_checkpoint(checkpoint_path, rows)
                gc.collect()
                percent = processed_count / total_jobs * 100 if total_jobs else 100.0
                print(
                    "进度："
                    f"{processed_count}/{total_jobs}（{percent:.1f}%）；"
                    f"高置信 {progress_counts.get('high_confidence_dead', 0)}，"
                    f"疑似 {progress_counts.get('review', 0)}，"
                    f"正常 {progress_counts.get('normal', 0)}，"
                    f"异常 {progress_counts.get('error', 0)}；"
                    f"checkpoint 已保存",
                    flush=True,
                )

        if checkpoint_path is not None:
            write_checkpoint(checkpoint_path, rows)
        print(
            "判断完成："
            f"{total_jobs}/{total_jobs}；"
            f"高置信 {progress_counts.get('high_confidence_dead', 0)}，"
            f"疑似 {progress_counts.get('review', 0)}，"
            f"正常 {progress_counts.get('normal', 0)}，"
            f"异常 {progress_counts.get('error', 0)}",
            flush=True,
        )

    df = pd.DataFrame(rows)
    df.attrs["results_by_cell"] = results_by_cell
    return df


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="检测 Excel 嵌入商品主图中的疑似死图底部状态条")
    parser.add_argument("input_xlsx", type=Path, help="输入 xlsx 文件路径")
    parser.add_argument("--output-dir", type=Path, default=Path("dead_image_output"), help="最终表格输出目录")
    parser.add_argument("--template-xlsx", type=Path, default=None, help="示例输出 xlsx；用于读取版式和前 N 行人工标注")
    parser.add_argument(
        "--correction-xlsx",
        type=Path,
        action="append",
        default=[],
        help="人工复核 xlsx；黄色死图判断单元格会作为正常图负样本。可重复传入多份。",
    )
    parser.add_argument(
        "--template-reference-dir",
        type=Path,
        default=None,
        help="死图模板参考图库目录；默认自动查找 template_reference，并把模板图案作为死图正样本训练",
    )
    parser.add_argument(
        "--disable-template-reference",
        action="store_true",
        help="不使用 template_reference 死图模板参考样本",
    )
    parser.add_argument("--calibration-rows", type=int, default=36, help="从示例输出中读取多少个数据行作为人工校准样本")
    parser.add_argument("--output-xlsx", type=Path, default=None, help="最终输出 Excel 文件路径")
    parser.add_argument(
        "--feedback-model",
        type=Path,
        default=Path("dead_image_output") / "models" / "current" / "dead_image_model_feedback.json",
        help="人工批注学习后的反馈模型数据文件；会自动加载并追加保存黄色误判样本",
    )
    parser.add_argument("--batch-size", type=int, default=500, help="每处理多少张图片保存一次 checkpoint")
    parser.add_argument("--checkpoint-jsonl", type=Path, default=None, help="阶段性处理明细 checkpoint；不填则输出到结果文件同名 jsonl")
    parser.add_argument("--sheet-name", default=None, help="工作表名称；不填则读取第一个/当前工作表")
    parser.add_argument("--id-column", default="商品ID", help="商品ID列名")
    parser.add_argument(
        "--image-columns",
        default=",".join(DEFAULT_IMAGE_COLUMNS),
        help="需要处理的图片列名，用英文逗号分隔",
    )

    # 常用调参项：如果误报多，可提高 high-confidence/review 阈值；
    # 如果漏报多，可降低阈值或扩大底部扫描区域。
    parser.add_argument("--high-confidence-threshold", type=float, default=0.90, help="高置信死图阈值；越高，填 1 越保守")
    parser.add_argument("--review-threshold", type=float, default=0.55, help="待复核/疑似死图阈值")
    parser.add_argument("--scan-start-frac", type=float, default=0.55, help="从图片高度的哪个比例开始扫描底部区域")
    parser.add_argument("--min-bar-height-frac", type=float, default=0.035, help="候选横条最小高度占比")
    parser.add_argument("--max-bar-height-frac", type=float, default=0.32, help="候选横条最大高度占比")
    parser.add_argument("--llm-enabled", action="store_true", help="????????????????????")
    parser.add_argument("--llm-api-url", default=None, help="OpenAI ??? chat/completions API ??")
    parser.add_argument("--llm-api-key", default=None, help="??? API Key???? DEAD_IMAGE_LLM_API_KEY ????")
    parser.add_argument("--llm-model", default=None, help="?????????? qwen-vl-max ? qwen3-vl")
    parser.add_argument("--llm-timeout", type=int, default=60, help="???????????????")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    image_columns = [item.strip() for item in args.image_columns.split(",") if item.strip()]
    config = DetectionConfig(
        scan_start_frac=args.scan_start_frac,
        min_bar_height_frac=args.min_bar_height_frac,
        max_bar_height_frac=args.max_bar_height_frac,
        high_confidence_threshold=args.high_confidence_threshold,
        review_threshold=args.review_threshold,
    )

    manual_labels: dict[tuple[int, str], int] | None = None
    judgment_cols: dict[str, int] = {}
    if args.template_xlsx:
        manual_labels, judgment_cols = read_template_layout(
            template_xlsx=args.template_xlsx,
            image_column_names=image_columns,
            calibration_rows=args.calibration_rows,
        )
        # 示例文件决定最终输出列；没有出现在示例中的图片列不写入版式输出。
        if judgment_cols:
            image_columns = [name for name in image_columns if name in judgment_cols]

    for correction_xlsx in args.correction_xlsx:
        correction_labels, correction_cols = read_yellow_corrections(
            correction_xlsx=correction_xlsx,
            image_column_names=image_columns,
        )
        manual_labels = manual_labels or {}
        manual_labels.update(correction_labels)
        if not judgment_cols and correction_cols:
            judgment_cols = correction_cols

    template_reference_vectors: list[np.ndarray] = []
    if not args.disable_template_reference:
        reference_dir = find_template_reference_dir(args.input_xlsx, args.template_reference_dir)
        template_reference_vectors = read_template_reference_vectors(reference_dir, config)

    feedback_vectors, feedback_labels, feedback_keys = load_feedback_samples(args.feedback_model)
    llm_config = llm_config_from_args(args)
    output_xlsx_path = args.output_xlsx or (args.output_dir / "死图判断结果.xlsx")
    checkpoint_path = args.checkpoint_jsonl or output_xlsx_path.with_suffix(".checkpoint.jsonl")

    df = process_workbook(
        input_xlsx=args.input_xlsx,
        sheet_name=args.sheet_name,
        id_column_name=args.id_column,
        image_column_names=image_columns,
        config=config,
        manual_labels=manual_labels,
        template_reference_vectors=template_reference_vectors,
        feedback_vectors=feedback_vectors,
        feedback_labels=feedback_labels,
        feedback_path=args.feedback_model,
        feedback_existing_keys=feedback_keys,
        batch_size=args.batch_size,
        checkpoint_path=checkpoint_path,
        llm_config=llm_config,
    )

    if args.template_xlsx and judgment_cols:
        write_template_output(
            template_xlsx=args.template_xlsx,
            output_xlsx=output_xlsx_path,
            results_by_cell=df.attrs.get("results_by_cell", {}),
            judgment_cols=judgment_cols,
        )
    else:
        output_xlsx_path.parent.mkdir(parents=True, exist_ok=True)
        df.to_excel(output_xlsx_path, index=False)
    style_final_workbook(output_xlsx_path)
    write_review_report(output_xlsx_path, df.to_dict("records"))

    total = len(df)
    high = int((df["判断级别"] == "high_confidence_dead").sum()) if total else 0
    review = int((df["判断级别"] == "review").sum()) if total else 0
    normal = int((df["判断级别"] == "normal").sum()) if total else 0

    print(f"处理完成：{total} 张图片")
    print(f"高置信死图：{high}，待复核：{review}，正常图：{normal}")
    print(f"最终表格：{output_xlsx_path}")
    print(f"checkpoint：{checkpoint_path}")
    print(f"二次审核明细：{output_xlsx_path.with_name(output_xlsx_path.stem + '_二次审核明细.xlsx')}")


if __name__ == "__main__":
    main()
