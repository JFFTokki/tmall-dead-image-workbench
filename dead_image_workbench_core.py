from __future__ import annotations

import re
from copy import copy
from dataclasses import dataclass
from io import BytesIO
from pathlib import Path

import openpyxl
from openpyxl.drawing.image import Image as XLImage


ROOT = Path(__file__).resolve().parent
DETECT_SCRIPT = ROOT / "死图判断源数据" / "detect_dead_images.py"
EXPORT_SCRIPT = ROOT / "maijsoft_export.py"
DEFAULT_TEMPLATE = ROOT / "死图判断源数据" / "export_469_20260706_135143_2932072_输出示例.xlsx"
DEFAULT_CORRECTION = ROOT / "optimized_result_template_labels_fixed - 批注.xlsx"
DEFAULT_HAR = ROOT / "tb.maijsoft.cn.har"
DEFAULT_INPUT = ROOT / "最新版死图.xlsx"
OUTPUT_ROOT = ROOT / "dead_image_output" / "workbench_runs"


def find_first_existing(*paths: Path) -> Path:
    for path in paths:
        if path.exists():
            return path
    return paths[0]


@dataclass
class RunOutputs:
    run_dir: Path
    main_xlsx: Path
    problem_xlsx: Path | None
    checkpoint_jsonl: Path
    log_file: Path


def create_problem_workbook(main_xlsx: Path) -> tuple[Path, int, int]:
    wb = openpyxl.load_workbook(main_xlsx)
    ws = wb.active
    total_rows = max(0, ws.max_row - 1)
    problem_rows: list[int] = []

    for row_idx in range(2, ws.max_row + 1):
        if row_has_problem(ws, row_idx):
            problem_rows.append(row_idx)

    problem_path = main_xlsx.with_name(main_xlsx.stem + "_问题结果表.xlsx")
    source_to_output_row = {source_row: index for index, source_row in enumerate(problem_rows, start=2)}
    out_wb = openpyxl.Workbook()
    out_ws = out_wb.active
    out_ws.title = "问题结果"

    for col in range(1, ws.max_column + 1):
        copy_cell(ws.cell(1, col), out_ws.cell(1, col))
        out_ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = ws.column_dimensions[
            openpyxl.utils.get_column_letter(col)
        ].width

    for out_row, source_row in enumerate(problem_rows, start=2):
        out_ws.row_dimensions[out_row].height = ws.row_dimensions[source_row].height
        for col in range(1, ws.max_column + 1):
            copy_cell(ws.cell(source_row, col), out_ws.cell(out_row, col))

    copy_images_for_problem_rows(ws, out_ws, source_to_output_row)
    out_ws.freeze_panes = "A2"
    problem_path.parent.mkdir(parents=True, exist_ok=True)
    out_wb.save(problem_path)
    out_wb.close()
    wb.close()
    return problem_path, total_rows, len(problem_rows)


def copy_images_for_problem_rows(
    source_ws: openpyxl.worksheet.worksheet.Worksheet,
    target_ws: openpyxl.worksheet.worksheet.Worksheet,
    source_to_output_row: dict[int, int],
) -> None:
    for image in getattr(source_ws, "_images", []):
        anchor = getattr(image, "anchor", None)
        marker = getattr(anchor, "_from", None)
        if marker is None:
            continue
        source_row = marker.row + 1
        target_row = source_to_output_row.get(source_row)
        if target_row is None:
            continue
        target_col = marker.col + 1
        try:
            copied = XLImage(BytesIO(image._data()))
        except Exception:
            continue
        copied.width = image.width
        copied.height = image.height
        copied.anchor = f"{openpyxl.utils.get_column_letter(target_col)}{target_row}"
        target_ws.add_image(copied)


def row_has_problem(ws: openpyxl.worksheet.worksheet.Worksheet, row_idx: int) -> bool:
    problem_values = {"1", "2", "是", "疑似", "review", "high_confidence_dead", "error"}
    for col in range(1, ws.max_column + 1):
        value = ws.cell(row_idx, col).value
        normalized = str(value).strip() if value is not None else ""
        if normalized in problem_values:
            return True
    return False


def copy_cell(source: openpyxl.cell.cell.Cell, target: openpyxl.cell.cell.Cell) -> None:
    target.value = source.value
    if source.has_style:
        target._style = copy(source._style)
    if source.number_format:
        target.number_format = source.number_format
    if source.alignment:
        target.alignment = copy(source.alignment)
    if source.fill:
        target.fill = copy(source.fill)
    if source.font:
        target.font = copy(source.font)
    if source.border:
        target.border = copy(source.border)


def zh(escaped: str) -> str:
    return escaped.encode("ascii").decode("unicode_escape")


def parse_progress(text: str) -> tuple[str, float] | None:
    percent_match = re.search(r"(\d+)\s*/\s*(\d+).*?(\d+(?:\.\d+)?)\s*%", text)
    if percent_match:
        done = int(percent_match.group(1))
        total = int(percent_match.group(2))
        raw_percent = float(percent_match.group(3))
        percent = min(94.0, max(10.0, raw_percent))
        counts = extract_progress_counts(text)
        suffix = zh("\\uff1b") + counts if counts else ""
        return f"{zh('\\u6b63\\u5728\\u5224\\u65ad\\uff1a')}{done}/{total}{zh('\\uff08')}{raw_percent:.1f}%{zh('\\uff09')}{suffix}", percent

    count_match = re.search(r"(\d+)\s*/\s*(\d+)", text)
    if count_match:
        done = int(count_match.group(1))
        total = max(1, int(count_match.group(2)))
        percent = min(94.0, max(10.0, done / total * 100))
        counts = extract_progress_counts(text)
        suffix = zh("\\uff1b") + counts if counts else ""
        return f"{zh('\\u6b63\\u5728\\u5224\\u65ad\\uff1a')}{done}/{total}{zh('\\uff08')}{percent:.1f}%{zh('\\uff09')}{suffix}", percent

    start_match = re.search(r"(\d+)", text) if zh("\\u5f00\\u59cb\\u5224\\u65ad") in text else None
    if start_match:
        return f"{zh('\\u5f00\\u59cb\\u5224\\u65ad\\uff1a\\u5171 ')}{start_match.group(1)}{zh(' \\u5f20\\u56fe\\u7247')}", 10.0
    return None


def extract_progress_counts(text: str) -> str:
    numbers = re.findall(r"\d+", text)
    if len(numbers) >= 6:
        return f"{zh('\\u9ad8\\u7f6e\\u4fe1 ')}{numbers[-4]}{zh('\\uff0c\\u7591\\u4f3c ')}{numbers[-3]}{zh('\\uff0c\\u6b63\\u5e38 ')}{numbers[-2]}{zh('\\uff0c\\u5f02\\u5e38 ')}{numbers[-1]}"
    return ""


def repair_mojibake(text: str) -> str:
    if not text:
        return text
    # Common Windows mojibake: UTF-8 bytes decoded as cp936/gbk first.
    for source_encoding in ("cp936", "gbk", "latin1"):
        try:
            repaired = text.encode(source_encoding, errors="strict").decode("utf-8", errors="strict")
        except UnicodeError:
            continue
        if mojibake_score(repaired) < mojibake_score(text):
            return repaired
    return text


def mojibake_score(text: str) -> int:
    suspicious = "???????????????"
    return sum(text.count(char) for char in suspicious) + text.count("?") * 3 + text.count("?")


def safe_stem(path: Path) -> str:
    stem = path.stem.strip() or "源数据"
    return re.sub(r'[<>:"/\\|?*]+', "_", stem)


def quote_arg(value: str) -> str:
    if not value or any(char.isspace() for char in value):
        return f'"{value}"'
    return value
